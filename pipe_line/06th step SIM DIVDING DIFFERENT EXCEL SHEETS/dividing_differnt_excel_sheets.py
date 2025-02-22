import os
import pandas as pd
from openpyxl import load_workbook
from unidecode import unidecode as uni
import re

# Define file paths
input_file_path = "example_monitoring_results.xlsx"
output_folder = "06th step SIM DIVIDING DIFFERENT EXCEL SHEETS/divided_different_excel_sheets/"

def process_file(file_path, output_folder):
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # Unmerge all cells
        merged_cells_copy = list(ws.merged_cells)
        for cell in merged_cells_copy:
            ws.unmerge_cells(str(cell))

        # Set A2 to "date"
        ws["A1"].value = None
        ws["A2"].value = None        
        ws["A2"].value = "date"

        # Save the modified workbook
        wb.save(file_path)

        # Load data into pandas DataFrame
        data = pd.read_excel(file_path)

        # Prepare the output filename
        selected_file_name = os.path.splitext(os.path.basename(file_path))[0]
        output_path = os.path.join(output_folder, f"{selected_file_name}.xlsx")

        # Process column names and create individual sheets
        new_column_names = data.columns
        updated_columns = [col for col in new_column_names if "Unnamed" not in col]

        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            for col in updated_columns:
                # Clean and prepare the sheet name
                sheet_name = uni(col)
                sheet_name = re.sub(r"[^\w\s]", "_", sheet_name)  # Replace non-word characters with "_"
                sheet_name = re.sub(r"\s+", "_", sheet_name)  # Replace spaces with "_"
                sheet_name = re.sub(r"_+", "_", sheet_name)  # Replace multiple "_" with a single "_"
                sheet_name = sheet_name.lower()  # Convert to lowercase
                sheet_name = sheet_name[:31]  # Trim to 31 characters (Excel limit)

                # Get the range of columns related to this data
                start_idx = new_column_names.get_loc(col)
                end_idx = start_idx + 1
                while end_idx < len(new_column_names) and "Unnamed" in new_column_names[end_idx]:
                    end_idx += 1
                column_data = data.iloc[:, start_idx:end_idx]

                # Filter and clean the data
                column_data = column_data.dropna(axis=1, how='all')  # Drop columns with all NaN values
                column_data = column_data.loc[:, column_data.apply(pd.Series.nunique) > 2]  # Drop columns with fewer than 3 unique values

                if len(column_data.columns) > 0:
                    column_data.insert(0, "date", data.iloc[:, 0])  # Add date column back

                    # Filter rows based on the first non-null date index
                    first_non_null_date_idx = column_data['date'].first_valid_index()
                    column_data = column_data.iloc[first_non_null_date_idx:]

                    # Write to Excel sheet
                    column_data.to_excel(writer, sheet_name=sheet_name, index=False, header=None)
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {e}")

def process_folder():
    process_file(input_file_path, output_folder)

# Main function
def main():
    process_folder()
    print("All processing completed successfully.")

if __name__ == "__main__":
    main()
